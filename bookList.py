import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox as msg
from openpyxl import Workbook, load_workbook

class BookList(tk.Toplevel):
    def __init__(self, db_connection):
        super().__init__()
        self.db_connection = db_connection
        self.db_cursor = self.db_connection.cursor()
        self.title("Book List")
        self.geometry("1000x500+600+300")
        self.resizable(width=False, height=False)
        #self.iconbitmap("python.ico")
        self.create_widgets()
        self.create_layout()
        self.protocol("WM_DELETE_WINDOW", self.close_window)
        self.load_books()

    def create_widgets(self):
        self.bookListlbl = ttk.Label(self, text="Book List", font=("Tahoma", 16), anchor="center")
        self.addBookBtn = ttk.Button(self, text="Add Book", command=self.add_book_window, bootstyle=SUCCESS)
        self.editBookBtn = ttk.Button(self, text="Edit Book", command=self.edit_book_window, bootstyle=WARNING)
        self.deleteBookBtn = ttk.Button(self, text="Delete Book", command=self.delete_book, bootstyle=DANGER)

        self.bookListTree = ttk.Treeview(self, columns=("ID", "Book Name", "Author", "Genre", "Status"), show="headings")

        self.bookListTree.heading("ID", text="ID")
        self.bookListTree.heading("Book Name", text="Book Name")
        self.bookListTree.heading("Author", text="Author")
        self.bookListTree.heading("Genre", text="Genre")
        self.bookListTree.heading("Status", text="Status")

        self.bookListTree.column("ID", width=50, anchor="center")
        self.bookListTree.column("Book Name", width=100, anchor="center")
        self.bookListTree.column("Author", width=100, anchor="center")
        self.bookListTree.column("Genre", width=150, anchor="center")
        self.bookListTree.column("Status", width=100, anchor="center")

        self.returnBtn = ttk.Button(self, text="Return Main Menu", command=self.close_window, bootstyle=SECONDARY)
        self.downloadSheetBtn = ttk.Button(self, text="Download Sheet", command=self.download_sheet, bootstyle=INFO)
        self.uploadSheetBtn = ttk.Button(self, text="Upload Sheet", command=self.upload_sheet, bootstyle=DARK)

    def create_layout(self):
        self.columnconfigure(0, weight=6)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)
        self.columnconfigure(3, weight=1)
        self.columnconfigure(4, weight=1)
        self.columnconfigure(5, weight=1)
        self.columnconfigure(6, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=9)

        self.bookListlbl.grid(row=0, column=0, pady=20)
        self.addBookBtn.grid(row = 0, column = 1, pady = 20, sticky = "nsew")
        self.editBookBtn.grid(row = 0, column = 2, pady = 20, sticky = "nsew")
        self.deleteBookBtn.grid(row = 0, column = 3, pady = 20, sticky = "nsew")
        self.downloadSheetBtn.grid(row = 0, column = 4, pady = 20, sticky = "nsew")
        self.uploadSheetBtn.grid(row = 0, column = 5, pady = 20, sticky = "nsew")
        self.returnBtn.grid(row = 0, column = 6, pady = 20, sticky = "nsew")
        self.bookListTree.grid(row = 1, column = 0, columnspan = 12, sticky = "nsew")

    def load_books(self):
        for row in self.bookListTree.get_children():
            self.bookListTree.delete(row)

        self.db_cursor.execute("SELECT * FROM books")
        books = self.db_cursor.fetchall()

        for book in books:
            self.bookListTree.insert("", "end", values=book)

    def add_book_window(self):
        book_add_win = tk.Toplevel(self)
        book_add_win.title("Add Book")
        book_add_win.geometry("250x200")
        book_add_win.resizable(width=False, height=False)
        book_add_win.grab_set()

        bookNameLbl = ttk.Label(book_add_win, text="Book Name:")
        bookNameEnt = ttk.Entry(book_add_win)

        bookAuthorLbl = ttk.Label(book_add_win, text="Author:")
        bookAuthorEnt = ttk.Entry(book_add_win)

        bookGenreLbl = ttk.Label(book_add_win, text="Genre:")
        bookGenreEnt = ttk.Entry(book_add_win)

        book_add_win.columnconfigure(0, weight=1)
        book_add_win.columnconfigure(1, weight=3)
        book_add_win.rowconfigure(0, weight=1)
        book_add_win.rowconfigure(1, weight=1)
        book_add_win.rowconfigure(2, weight=1)
        book_add_win.rowconfigure(3, weight=1)

        bookNameLbl.grid(row=0, column=0, pady=5, sticky="e")
        bookNameEnt.grid(row=0, column=1, pady=5, sticky="news")
        bookAuthorLbl.grid(row=1, column=0, pady=5, sticky="e")
        bookAuthorEnt.grid(row=1, column=1, pady=5, sticky="news")
        bookGenreLbl.grid(row=2, column=0, pady=5, sticky="e")
        bookGenreEnt.grid(row=2, column=1, pady=5, sticky="news")

        bookAddBtn = ttk.Button(book_add_win, text="Add",
                                  command=lambda: self.add_book(bookNameEnt.get(), bookAuthorEnt.get(),
                                                                  bookGenreEnt.get(),book_add_win))

        bookAddBtn.grid(row=3, column=0, columnspan=2, pady=10)

    def add_book(self, book_name, book_author, book_genre, window):
        try:
            if len(book_name) == 0 or len(book_author) == 0 or len(book_genre) == 0:
                msg.showerror("Error", "Please fill out all fields in the form.")
                return

            self.db_cursor.execute("INSERT INTO books (book_name, book_author, book_genre) VALUES (?, ?, ?)",
                                   (book_name, book_author, book_genre))
            self.db_connection.commit()
            self.load_books()
            window.destroy()
        except Exception as e:
            print("Error adding book:", e)
            msg.showerror("Error", "Unable to add book.")

    def edit_book_window(self):
        selected_item = self.bookListTree.selection()
        if not selected_item:
            msg.showerror("Error", "Please select a member to edit.")
            return

        book_id, book_name, book_author, book_genre, _ = self.bookListTree.item(selected_item, 'values')
        book_edit_win = tk.Toplevel(self)
        book_edit_win.title("Edit Book")
        book_edit_win.geometry("250x200")
        book_edit_win.resizable(width=False, height=False)
        book_edit_win.grab_set()

        bookNameLbl = ttk.Label(book_edit_win, text="Book Name:")
        bookNameEnt = ttk.Entry(book_edit_win)
        bookNameEnt.insert(0, book_name)

        bookAuthorLbl = ttk.Label(book_edit_win, text="Author:")
        bookAuthorEnt = ttk.Entry(book_edit_win)
        bookAuthorEnt.insert(0, book_author)

        bookGenreLbl = ttk.Label(book_edit_win, text="Genre:")
        bookGenreEnt = ttk.Entry(book_edit_win)
        bookGenreEnt.insert(0, book_genre)

        book_edit_win.columnconfigure(0, weight=1)
        book_edit_win.columnconfigure(1, weight=3)
        book_edit_win.rowconfigure(0, weight=1)
        book_edit_win.rowconfigure(1, weight=1)
        book_edit_win.rowconfigure(2, weight=1)
        book_edit_win.rowconfigure(3, weight=1)

        bookNameLbl.grid(row=0, column=0, pady=5, sticky="e")
        bookNameEnt.grid(row=0, column=1, pady=5, sticky="news")
        bookAuthorLbl.grid(row=1, column=0, pady=5, sticky="e")
        bookAuthorEnt.grid(row=1, column=1, pady=5, sticky="news")
        bookGenreLbl.grid(row=2, column=0, pady=5, sticky="e")
        bookGenreEnt.grid(row=2, column=1, pady=5, sticky="news")

        bookEditBtn = ttk.Button(book_edit_win, text="Edit",
                                  command=lambda: self.edit_book(book_id,bookNameEnt.get(), bookAuthorEnt.get(),
                                                                  bookGenreEnt.get(),book_edit_win))

        bookEditBtn.grid(row=3, column=0, columnspan=2, pady=10)

    def edit_book(self, book_id, book_name, book_author, book_genre, window):
        try:
            if len(book_name) == 0 or len(book_author) == 0 or len(book_genre) == 0:
                msg.showerror("Error", "Please fill out all fields in the form.")
                return

            self.db_cursor.execute("UPDATE books SET book_name = ?, book_author = ?, book_genre = ? WHERE id = ?",
                                    (book_name, book_author, book_genre, book_id)
            )
            self.db_connection.commit()
            self.load_books()
            window.destroy()
        except Exception as e:
            print("Error editing book:", e)
            msg.showerror("Error", "Unable to edit book.")

    def delete_book(self):
        selected_item = self.bookListTree.selection()
        if not selected_item:
            msg.showerror("Error", "Please select a book to delete.")
            return

        book_id = self.bookListTree.item(selected_item, 'values')[0]
        if msg.askyesno("Delete Book", "Are you sure you want to delete this book?"):
            try:
                self.db_cursor.execute("DELETE FROM books WHERE id = ?", (book_id,))
                self.db_connection.commit()
                self.load_books()
            except Exception as e:
                print("Error deleting book:", e)
                msg.showerror("Error", "Unable to delete book.")

    def download_sheet(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Books"

            headers = ["ID", "Title", "Author", "Genre", "Status"]
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            self.db_cursor.execute("SELECT * FROM books")
            books = self.db_cursor.fetchall()

            for row_num, book in enumerate(books, start=2):
                for col_num, value in enumerate(book, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save("books.xlsx")
            msg.showinfo("Success", "Books exported to books.xlsx")
        except Exception as e:
            print("Error exporting books:", e)
            msg.showerror("Error", "Unable to export books.")

    def upload_sheet(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel Files", "*.xlsx *.xls")]
            )

            if not file_path:
                msg.showinfo("Cancelled", "No file selected.")
                return

            workbook = load_workbook(file_path)
            sheet = workbook.active

            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            if not rows:
                msg.showinfo("Info", "The selected file is empty or not formatted properly.")
                return

            for row in rows:
                try:
                    self.db_cursor.execute(
                        "INSERT INTO books (book_name, book_author, book_genre, book_status) VALUES (?, ?, ?, ?)",
                        row[1:]
                    )
                except Exception as e:
                    print(f"Error inserting row {row}: {e}")

            # Commit the changes
            self.db_connection.commit()
            self.load_books()
            msg.showinfo("Success", "Books imported successfully from the Excel file.")
        except Exception as e:
            print("Error uploading books:", e)
            msg.showerror("Error", "Unable to upload books.")

    def close_window(self):
        self.destroy()
        self.master.deiconify()