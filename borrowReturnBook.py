import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox as msg
from openpyxl import Workbook, load_workbook

class BorrowReturnBook(tk.Toplevel):
    def __init__(self, db_connection):
        super().__init__()
        self.db_connection = db_connection
        self.db_cursor = self.db_connection.cursor()
        self.title("Borrow/Return Book")
        self.geometry("1000x500+600+300")
        self.resizable(width=False, height=False)
        #self.iconbitmap("python.ico")
        self.create_widgets()
        self.create_layout()
        self.protocol("WM_DELETE_WINDOW", self.close_window)
        self.load_borrow_return_books()

    def create_widgets(self):
        self.borrowReturnBooklbl = ttk.Label(self, text="Borrow/Return Book", font=("Tahoma", 16), anchor="center")
        self.borrowBookBtn = ttk.Button(self, text="Borrow Book", command=self.borrow_book_window, bootstyle=SUCCESS)
        self.returnBookBtn = ttk.Button(self, text="Return Book", command=self.return_book, bootstyle=DANGER)
        self.borrowReturnBooktTree = ttk.Treeview(self, columns=("ID", "Member ID", "Book ID", "Borrow Date", "Return Date"), show="headings")

        self.borrowReturnBooktTree.heading("ID", text="ID")
        self.borrowReturnBooktTree.heading("Member ID", text="Member ID")
        self.borrowReturnBooktTree.heading("Book ID", text="Book ID")
        self.borrowReturnBooktTree.heading("Borrow Date", text="Borrow Date")
        self.borrowReturnBooktTree.heading("Return Date", text="Return Date")

        self.borrowReturnBooktTree.column("ID", width=50, anchor="center")
        self.borrowReturnBooktTree.column("Member ID", width=100, anchor="center")
        self.borrowReturnBooktTree.column("Book ID", width=100, anchor="center")
        self.borrowReturnBooktTree.column("Borrow Date", width=150, anchor="center")
        self.borrowReturnBooktTree.column("Return Date", width=100, anchor="center")

        self.returnBtn = ttk.Button(self, text="Return Main Menu", command=self.close_window, bootstyle=SECONDARY)
        self.downloadSheetBtn = ttk.Button(self, text="Download Sheet", command=self.download_sheet, bootstyle=INFO)
        self.uploadSheetBtn = ttk.Button(self, text="Upload Sheet", command=self.upload_sheet, bootstyle=DARK)

    def create_layout(self):
        self.columnconfigure(0, weight=5)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)
        self.columnconfigure(3, weight=1)
        self.columnconfigure(4, weight=1)
        self.columnconfigure(5, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=9)

        self.borrowReturnBooklbl.grid(row=0, column=0, pady=20)
        self.borrowBookBtn.grid(row = 0, column = 1, pady = 20, sticky = "nsew")
        self.returnBookBtn.grid(row = 0, column = 2, pady = 20, sticky = "nsew")
        self.downloadSheetBtn.grid(row = 0, column = 3, pady = 20, sticky = "nsew")
        self.uploadSheetBtn.grid(row = 0, column = 4, pady = 20, sticky = "nsew")
        self.returnBtn.grid(row = 0, column = 5, pady = 20, sticky = "nsew")
        self.borrowReturnBooktTree.grid(row = 1, column = 0, columnspan = 10, sticky = "nsew")

    def load_borrow_return_books(self):
        for row in self.borrowReturnBooktTree.get_children():
            self.borrowReturnBooktTree.delete(row)

        self.db_cursor.execute("SELECT * FROM borrow_records")
        borrow_records = self.db_cursor.fetchall()

        for record in borrow_records:
            self.borrowReturnBooktTree.insert("", "end", values=record)

    def borrow_book_window(self):
        borrow_book_window = tk.Toplevel(self)
        borrow_book_window.title("Borrow Book")
        borrow_book_window.geometry("250x200")
        borrow_book_window.resizable(width=False, height=False)
        borrow_book_window.grab_set()

        borrowBookIdlbl = ttk.Label(borrow_book_window, text="Book ID:")
        self.db_cursor.execute("SELECT id, book_name FROM books WHERE book_status = 'Available'")
        book_ids = [f"{row[0]} - {row[1]}" for row in self.db_cursor.fetchall()]
        borrowBookIdCombo = ttk.Combobox(borrow_book_window, values=book_ids)

        borrowMemberIdlbl = ttk.Label(borrow_book_window, text="Member ID:")
        self.db_cursor.execute("SELECT id, name, surname FROM members")
        member_ids = [f"{row[0]} - {row[1]} {row[2]}" for row in self.db_cursor.fetchall()]
        borrowMemberIdCombo = ttk.Combobox(borrow_book_window, values=member_ids)

        borrow_book_window.columnconfigure(0, weight=1)
        borrow_book_window.columnconfigure(1, weight=2)
        borrow_book_window.rowconfigure(0, weight=1)
        borrow_book_window.rowconfigure(1, weight=1)
        borrow_book_window.rowconfigure(2, weight=1)

        borrowBookIdlbl.grid(row=0, column=0, pady=10)
        borrowBookIdCombo.grid(row=0, column=1, pady=10)
        borrowMemberIdlbl.grid(row=1, column=0, pady=10)
        borrowMemberIdCombo.grid(row=1, column=1, pady=10)

        borrowBookBtn = ttk.Button(borrow_book_window, text="Borrow Book", command=lambda: self.borrow_book(borrowBookIdCombo.get(), borrowMemberIdCombo.get(), borrow_book_window))

        borrowBookBtn.grid(row=2, column=0, columnspan=2, pady=10)

    def borrow_book(self, book_id, member_id, window):
        try:
            book_id = book_id.split(" - ")[0]
            member_id = member_id.split(" - ")[0]

            self.db_cursor.execute("INSERT INTO borrow_records (book_id, member_id, borrow_date) VALUES (?, ?, datetime('now'))", (book_id, member_id))
            self.db_connection.commit()
            self.db_cursor.execute("UPDATE books SET book_status = 'Borrowed' WHERE id = ?", (book_id,))
            self.db_connection.commit()
            self.load_borrow_return_books()
            window.destroy()
        except Exception as e:
            print("Error barrowing book:", e)
            msg.showerror("Error", "Error borrowing book.")

    def return_book(self):
        try:
            selected_item = self.borrowReturnBooktTree.selection()
            if not selected_item:
                msg.showerror("Error", "Please select a record to edit.")
                return

            id, _, book_id, _, _ = self.borrowReturnBooktTree.item(selected_item, 'values')

            self.db_cursor.execute("UPDATE borrow_records SET return_date = datetime('now') WHERE id = ?", (id,))
            self.db_connection.commit()
            self.db_cursor.execute("UPDATE books SET book_status = 'Available' WHERE id = ?", (book_id,))
            self.db_connection.commit()
            self.load_borrow_return_books()
        except Exception as e:
            print("Error returning book:", e)
            msg.showerror("Error", "Error returning book.")

    def download_sheet(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Books"

            headers = ["ID", "Member ID", "Book ID", "Borrow Date", "Return Date"]
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            self.db_cursor.execute("SELECT * FROM borrow_records")
            books = self.db_cursor.fetchall()

            for row_num, book in enumerate(books, start=2):
                for col_num, value in enumerate(book, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save("borrowRecords.xlsx")
            msg.showinfo("Success", "Borrow Records exported to borrowRecords.xlsx")
        except Exception as e:
            print("Error exporting Borrow Records:", e)
            msg.showerror("Error", "Unable to export Borrow Records.")

    def upload_sheet(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel Files", "*.xlsx *.xls")]
            )

            if not file_path:
                msg.showinfo("Cancelled", "No file selected.")
                return

            workbook = load_workbook(file_path)
            sheet = workbook.active

            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            if not rows:
                msg.showinfo("Info", "The selected file is empty or not formatted properly.")
                return

            for row in rows:
                try:
                    self.db_cursor.execute(
                        "INSERT INTO borrow_records (member_id, book_id, borrow_date, return_date) VALUES (?, ?, ?, ?)",
                        row[1:]
                    )
                except Exception as e:
                    print(f"Error inserting row {row}: {e}")

            self.db_connection.commit()
            self.load_borrow_return_books()
            msg.showinfo("Success", "Borrow Records imported successfully from the Excel file.")
        except Exception as e:
            print("Error uploading Borrow Records:", e)
            msg.showerror("Error", "Unable to upload Borrow Records.")

    def close_window(self):
        self.destroy()
        self.master.deiconify()