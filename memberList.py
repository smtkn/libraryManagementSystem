import re
import tkinter as tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox as msg
from openpyxl import Workbook, load_workbook


class MemberList(tk.Toplevel):
    def __init__(self, db_connection):
        super().__init__()
        self.db_connection = db_connection
        self.db_cursor = self.db_connection.cursor()
        self.title("Member List")
        self.geometry("1000x500+600+300")
        self.resizable(width=False, height=False)
        self.create_widgets()
        self.create_layout()
        self.protocol("WM_DELETE_WINDOW", self.close_window)
        self.load_members()

    def create_widgets(self):
        self.memberListlbl = ttk.Label(self, text="Member List", font=("Tahoma", 16), anchor="center")
        self.addMemberBtn = ttk.Button(self, text="Add Member", command=self.add_member_window, width=20, bootstyle=SUCCESS)
        self.editMemberBtn = ttk.Button(self, text="Edit Member", command=self.edit_member_window, width=20, bootstyle=WARNING)
        self.deleteMemberBtn = ttk.Button(self, text="Delete Member", command=self.delete_member, width=20, bootstyle=DANGER)

        self.memberListTree = ttk.Treeview(self, columns=("ID", "Name", "Surname", "Email", "Phone"), show="headings")
        self.memberListTree.heading("ID", text="ID")
        self.memberListTree.heading("Name", text="Name")
        self.memberListTree.heading("Surname", text="Surname")
        self.memberListTree.heading("Email", text="Email")
        self.memberListTree.heading("Phone", text="Phone")

        self.memberListTree.column("ID", width=50, anchor="center")
        self.memberListTree.column("Name", width=150, anchor="center")
        self.memberListTree.column("Surname", width=150, anchor="center")
        self.memberListTree.column("Email", width=200, anchor="center")
        self.memberListTree.column("Phone", width=100, anchor="center")

        self.returnBtn = ttk.Button(self, text="Return Main Menu", command=self.close_window, width=20, bootstyle=SECONDARY)
        self.downloadSheetBtn = ttk.Button(self, text="Download Sheet", command=self.download_sheet, bootstyle=INFO)
        self.uploadSheetBtn = ttk.Button(self, text="Upload Sheet", command=self.upload_sheet, bootstyle=DARK)

    def create_layout(self):
        self.columnconfigure(0, weight=4)
        self.columnconfigure(1, weight=1)
        self.columnconfigure(2, weight=1)
        self.columnconfigure(3, weight=1)
        self.columnconfigure(4, weight=1)
        self.columnconfigure(5, weight=1)
        self.columnconfigure(6, weight=1)
        self.rowconfigure(0, weight=1)
        self.rowconfigure(1, weight=9)

        self.memberListlbl.grid(row=0, column=0, pady=20)
        self.addMemberBtn.grid(row=0, column=1, pady=20, sticky="nsew")
        self.editMemberBtn.grid(row=0, column=2, pady=20, sticky="nsew")
        self.deleteMemberBtn.grid(row=0, column=3, pady=20, sticky="nsew")
        self.downloadSheetBtn.grid(row=0, column=4, pady=20, sticky="nsew")
        self.uploadSheetBtn.grid(row=0, column=5, pady=20, sticky="nsew")
        self.returnBtn.grid(row=0, column=6, pady=20, sticky="nsew")
        self.memberListTree.grid(row=1, column=0, columnspan=8, sticky="nsew")

    def load_members(self):
        for row in self.memberListTree.get_children():
            self.memberListTree.delete(row)

        self.db_cursor.execute("SELECT * FROM members")
        members = self.db_cursor.fetchall()

        for member in members:
            self.memberListTree.insert("", "end", values=member)

    def add_member_window(self):
        member_add_win = tk.Toplevel(self)
        member_add_win.title("Add Member")
        member_add_win.geometry("250x200")
        member_add_win.resizable(width=False, height=False)
        member_add_win.grab_set()

        memberNameLbl = ttk.Label(member_add_win, text="Name:")
        memberNameEnt = ttk.Entry(member_add_win)

        memberSurnameLbl = ttk.Label(member_add_win, text="Surname:")
        memberSurnameEnt = ttk.Entry(member_add_win)

        memberEmailLbl = ttk.Label(member_add_win, text="Email:")
        memberEmailEnt = ttk.Entry(member_add_win)

        memberPhoneLbl = ttk.Label(member_add_win, text="Phone:")
        memberPhoneEnt = ttk.Entry(member_add_win)

        member_add_win.columnconfigure(0, weight=3)
        member_add_win.columnconfigure(1, weight=1)
        member_add_win.rowconfigure(0, weight=1)
        member_add_win.rowconfigure(1, weight=1)
        member_add_win.rowconfigure(2, weight=1)
        member_add_win.rowconfigure(3, weight=1)
        member_add_win.rowconfigure(4, weight=1)

        memberNameLbl.grid(row=0, column=0, pady=5, sticky="e")
        memberNameEnt.grid(row=0, column=1, pady=5, sticky="news")
        memberSurnameLbl.grid(row=1, column=0, pady=5, sticky="e")
        memberSurnameEnt.grid(row=1, column=1, pady=5, sticky="news")
        memberEmailLbl.grid(row=2, column=0, pady=5, sticky="e")
        memberEmailEnt.grid(row=2, column=1, pady=5, sticky="news")
        memberPhoneLbl.grid(row=3, column=0, pady=5, sticky="e")
        memberPhoneEnt.grid(row=3, column=1, pady=5, sticky="news")

        memberAddBtn = ttk.Button(member_add_win, text="Add",
                   command=lambda: self.add_member(memberNameEnt.get(), memberSurnameEnt.get(), memberEmailEnt.get(), memberPhoneEnt.get(), member_add_win))

        memberAddBtn.grid(row=4, column=0, columnspan=2, pady=10)

    def add_member(self, name, surname, email, phone, window):
        try:
            if len(name) == 0 or len(surname) == 0 or len(email) == 0 or len(phone) == 0:
                msg.showerror("Error", "Please fill out all fields in the form.")
                return

            email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
            if not re.match(email_regex, email):
                msg.showerror("Error", "Invalid email address.")
                return

            if not phone.isdigit():
                msg.showerror("Error", "Phone number must be a number.")
                return

            if len(phone) < 11 or not phone.startswith("05"):
                msg.showerror("Error", "Phone number must be at least 11 (05XXXXXXXXX) digits.")
                return

            self.db_cursor.execute("INSERT INTO members (name, surname, email, phone) VALUES (?, ?, ?, ?)", (name, surname, email, phone))
            self.db_connection.commit()
            self.load_members()
            window.destroy()
        except Exception as e:
            print("Error adding member:", e)
            msg.showerror("Error", "Unable to add member.")

    def edit_member_window(self):
        selected_item = self.memberListTree.selection()
        if not selected_item:
            msg.showerror("Error", "Please select a member to edit.")
            return

        member_id, name, surname, email, phone = self.memberListTree.item(selected_item, 'values')
        member_edit_win = tk.Toplevel(self)
        member_edit_win.title("Add Member")
        member_edit_win.geometry("250x200")
        member_edit_win.resizable(width=False, height=False)
        member_edit_win.grab_set()

        memberNameLbl = ttk.Label(member_edit_win, text="Name:")
        memberNameEnt = ttk.Entry(member_edit_win)
        memberNameEnt.insert(0, name)

        memberSurnameLbl = ttk.Label(member_edit_win, text="Surname:")
        memberSurnameEnt = ttk.Entry(member_edit_win)
        memberSurnameEnt.insert(0, surname)

        memberEmailLbl = ttk.Label(member_edit_win, text="Email:")
        memberEmailEnt = ttk.Entry(member_edit_win)
        memberEmailEnt.insert(0, email)

        memberPhoneLbl = ttk.Label(member_edit_win, text="Phone:")
        memberPhoneEnt = ttk.Entry(member_edit_win)
        memberPhoneEnt.insert(0, phone)

        member_edit_win.columnconfigure(0, weight=1)
        member_edit_win.columnconfigure(1, weight=3)
        member_edit_win.rowconfigure(0, weight=1)
        member_edit_win.rowconfigure(1, weight=1)
        member_edit_win.rowconfigure(2, weight=1)
        member_edit_win.rowconfigure(3, weight=1)
        member_edit_win.rowconfigure(4, weight=1)
        member_edit_win.rowconfigure(5, weight=1)

        memberNameLbl.grid(row=0, column=0, pady=5, sticky="e")
        memberNameEnt.grid(row=0, column=1, pady=5, sticky="news")
        memberSurnameLbl.grid(row=1, column=0, pady=5, sticky="e")
        memberSurnameEnt.grid(row=1, column=1, pady=5, sticky="news")
        memberEmailLbl.grid(row=2, column=0, pady=5, sticky="e")
        memberEmailEnt.grid(row=2, column=1, pady=5, sticky="news")
        memberPhoneLbl.grid(row=3, column=0, pady=5, sticky="e")
        memberPhoneEnt.grid(row=3, column=1, pady=5, sticky="news")

        memberEditBtn = ttk.Button(member_edit_win, text="Edit",
                                  command=lambda: self.update_member(member_id, memberNameEnt.get(), memberSurnameEnt.get(), memberEmailEnt.get(), memberPhoneEnt.get(), member_edit_win))

        memberEditBtn.grid(row=5, column=0, columnspan=2, pady=10)

    def update_member(self, member_id, name, surname, email, phone, window):
        try:
            if len(name) == 0 or len(surname) == 0 or len(email) == 0 or len(phone) == 0:
                msg.showerror("Error", "Please fill out all fields in the form.")
                return

            email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
            if not re.match(email_regex, email):
                msg.showerror("Error", "Invalid email address.")
                return

            if not phone.isdigit():
                msg.showerror("Error", "Phone number must be a number.")
                return

            if len(phone) < 11 or not phone.startswith("05"):
                msg.showerror("Error", "Phone number must be at least 11 (05XXXXXXXXX) digits.")
                return

            self.db_cursor.execute("UPDATE members SET name = ?, surname = ?, email = ?, phone = ? WHERE id = ?",
                                   (name, surname, email, phone, member_id))
            self.db_connection.commit()
            self.load_members()
            window.destroy()
        except Exception as e:
            print("Error updating member:", e)
            msg.showerror("Error", "Unable to update member.")

    def delete_member(self):
        selected_item = self.memberListTree.selection()
        if not selected_item:
            msg.showerror("Error", "Please select a member to delete.")
            return

        member_id = self.memberListTree.item(selected_item, 'values')[0]
        if msg.askyesno("Delete Member", "Are you sure you want to delete this member?"):
            try:
                self.db_cursor.execute("DELETE FROM members WHERE id = ?", (member_id,))
                self.db_connection.commit()
                self.load_members()
            except Exception as e:
                print("Error deleting member:", e)
                msg.showerror("Error", "Unable to delete member.")

    def download_sheet(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Books"

            headers = ["ID", "Member Name", "Member Surname", "Member Email", "Member Phone"]
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)

            self.db_cursor.execute("SELECT * FROM members")
            books = self.db_cursor.fetchall()

            for row_num, book in enumerate(books, start=2):
                for col_num, value in enumerate(book, start=1):
                    sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save("members.xlsx")
            msg.showinfo("Success", "Members exported to members.xlsx")
        except Exception as e:
            print("Error exporting Borrow Records:", e)
            msg.showerror("Error", "Unable to export Members.")

    def upload_sheet(self):
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel File",
                filetypes=[("Excel Files", "*.xlsx *.xls")]
            )

            if not file_path:
                msg.showinfo("Cancelled", "No file selected.")
                return

            workbook = load_workbook(file_path)
            sheet = workbook.active

            rows = list(sheet.iter_rows(min_row=2, values_only=True))

            if not rows:
                msg.showinfo("Info", "The selected file is empty or not formatted properly.")
                return

            for row in rows:
                try:
                    self.db_cursor.execute(
                        "INSERT INTO members (name, surname, email, phone) VALUES (?, ?, ?, ?)",
                        row[1:]
                    )
                except Exception as e:
                    print(f"Error inserting row {row}: {e}")

            self.db_connection.commit()
            self.load_members()
            msg.showinfo("Success", "Members imported successfully from the Excel file.")
        except Exception as e:
            print("Error uploading Members:", e)
            msg.showerror("Error", "Unable to upload Members.")

    def close_window(self):
        self.destroy()
        self.master.deiconify()
